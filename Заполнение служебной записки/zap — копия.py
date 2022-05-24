import fnmatch
import os

import openpyxl
from PyPDF2 import PdfReader
from openpyxl.styles import Border, Side, Font, Alignment

excel_path = os.getcwd()
listOfFiles = os.listdir('.')

pdf_files = [f for f in listOfFiles if f.endswith('.pdf')]
pdf_files.sort()

r = p = k = 11
# p = 11
l = 0
number = 1
# k = 11
ob_list = []
naim_list = []
count_list = []

for name in pdf_files:
    global name1
    name2 = name.split('.pdf')
    name_len = len(pdf_files)
    name1 = name2.pop(0)
    name = name1.split(' ')

    global ob
    ob = name.pop(0)
    ob = ob.split('БТЛИ.')
    ob = ob.pop(1)

    global naim
    naim = ' '.join(name)
    ob_list.append(ob)  # пополнить список обозначений
    naim_list.append(naim)  # пополнить список наименований
    # чтение PDF
    reader = PdfReader(pdf_files[l])
    number_of_pages = len(reader.pages)
    page = reader.pages[0]

    count_list.append(number_of_pages)
    l += 1

wb = openpyxl.load_workbook("Служебная записка на обработку и размножение чертежей.xlsx")
sheets = wb.sheetnames
sheet_ranges = wb['Лист1']
ws = wb.active
ws.title = 'Лист1'

string_number = len(ob_list)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

font = Font(name='Times New Roman',
            size=12,
            color='FF000000',
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False)

alignment = Alignment(horizontal='center',
                      vertical='center',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)

for row in ob_list:
    ws.cell(row=r, column=10).value = row
    ws.cell(row=r, column=10).border = thin_border
    ws.cell(row=r, column=10).font = font
    ws.cell(row=r, column=10).alignment = alignment
    r += 1
    ws.insert_rows(r + 1)
for row in naim_list:
    ws.cell(row=p, column=11).value = row
    ws.cell(row=p, column=11).border = thin_border
    ws.cell(row=p, column=11).font = font
    ws.cell(row=p, column=11).alignment = alignment
    p += 1
for row in count_list:
    ws.cell(row=k, column=12).value = row
    ws.cell(row=k, column=12).font = font
    ws.cell(row=k, column=12).alignment = alignment
    ws.cell(row=k, column=1).value = number
    ws.cell(row=k, column=1).font = font
    ws.cell(row=k, column=1).alignment = alignment
    ws.cell(row=k, column=12).border = thin_border
    ws.cell(row=k, column=1).border = thin_border
    ws.cell(row=k, column=6).border = thin_border
    ws.cell(row=k, column=7).border = thin_border
    ws.cell(row=k, column=8).border = thin_border
    ws.cell(row=k, column=9).border = thin_border
    ws.cell(row=k, column=13).border = thin_border
    k += 1
    number += 1

ws.merge_cells(start_row=11, start_column=2, end_column=2, end_row=k - 1)
ws.merge_cells(start_row=11, start_column=3, end_column=3, end_row=k - 1)
ws.merge_cells(start_row=11, start_column=4, end_column=4, end_row=k - 1)
ws.merge_cells(start_row=11, start_column=5, end_column=5, end_row=k - 1)

wb.save('Служебная записка на обработку и размножение чертежей.xlsx')

print(count_list)
print(ob_list)
print(naim_list)
