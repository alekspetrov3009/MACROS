import fnmatch
import os

import PyPDF2
import openpyxl
from PyPDF2 import PdfReader
from openpyxl.styles import Border, Side, Font, Alignment

excel_path = os.getcwd()
listOfFiles = os.listdir('.')

pdf_files = [f for f in listOfFiles if f.endswith('.pdf')]
pdf_files.sort()

pdf_path = []
for root, dirs, files in os.walk(excel_path):
    for file in files:
        if not file.endswith(".pdf"):
            continue
        pdf_path.append(os.path.join(root, file))
        print(pdf_path)

r = p = k = 11
# p = 11
l = 0
number = 1
# k = 11
ob_list = []
naim_list = []
count_list = []

for name in pdf_files:
    global ob
    global naim
    # global name
    global sb
    name = name.split('.pdf')  # отделяем формат от имени файла
    name1 = name.pop(0)  # выбираем первый элемент из списка
    name = name1.split(' ')  # отделяем пробелы
    ob = name.pop(0)  # выбираем в качестве обозначения первый элемент списка
    ob = ob.split('БТЛИ.')  # отделяем от обозначения БТЛИ
    ob = ob.pop(1)  # выбираем второй элемент списка
    if 'СБ' in name:  # проверка на присутствие СБ в названии
        # sb = name.pop(1)
        sb = 'СБ'
        ob = ob + sb
        naim = name.pop(1)
    else:
        sb = ' '
        ob = ob
        naim = name.pop(0)
    count_big_naim = len(sb)
    if count_big_naim > 1:
        global naim2
        naim2 = " ".join(name)
        # naim3 = naim2.split('СБ ')
        # naim = [' '.join(naim3)]
    ob_list.append(ob)  # пополнить список обозначений
    naim_list.append(naim)  # пополнить список наименований

    # чтение PDF
    reader = PdfReader(pdf_files[l])
    number_of_pages = len(reader.pages)
    page = reader.pages[0]

    count_list.append(number_of_pages)
    l += 1
    # чтение размеров листа
    sizes = []
    pdf = 0
    # print(pdf_str)
    for pdf in pdf_path:
        with open(pdf, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfFileReader(pdf_file)
            i = 0
            while True:
                try:
                    box = pdf_reader.getPage(i)
                    width = int(float(box.mediaBox.getWidth()) * 0.3527)
                    height = int(float(box.mediaBox.getHeight()) * 0.3527)
                    sizes.append([height, width])
                    i += 1
                except Exception:
                    break

print(sizes)
# wb = openpyxl.load_workbook("Служебная записка на обработку и размножение чертежей.xlsx")
# sheets = wb.sheetnames
# sheet_ranges = wb['Лист1']
# ws = wb.active
# ws.title = 'Лист1'
#
# string_number = len(ob_list)
#
# thin_border = Border(left=Side(style='thin'),
#                      right=Side(style='thin'),
#                      top=Side(style='thin'),
#                      bottom=Side(style='thin'))
#
# font = Font(name='Times New Roman',
#             size=12,
#             color='FF000000',
#             bold=False,
#             italic=False,
#             vertAlign=None,
#             underline='none',
#             strike=False)
#
# alignment = Alignment(horizontal='center',
#                       vertical='center',
#                       text_rotation=0,
#                       wrap_text=False,
#                       shrink_to_fit=False,
#                       indent=0)
#
# for row in ob_list:
#     ws.cell(row=r, column=10).value = row
#     ws.cell(row=r, column=10).border = thin_border
#     ws.cell(row=r, column=10).font = font
#     ws.cell(row=r, column=10).alignment = alignment
#     r += 1
#     ws.insert_rows(r + 1)
# for row in naim_list:
#     ws.cell(row=p, column=11).value = row
#     ws.cell(row=p, column=11).border = thin_border
#     ws.cell(row=p, column=11).font = font
#     ws.cell(row=p, column=11).alignment = alignment
#     p += 1
# for row in count_list:
#     ws.cell(row=k, column=12).value = row
#     ws.cell(row=k, column=12).font = font
#     ws.cell(row=k, column=12).alignment = alignment
#     ws.cell(row=k, column=1).value = number
#     ws.cell(row=k, column=1).font = font
#     ws.cell(row=k, column=1).alignment = alignment
#     ws.cell(row=k, column=12).border = thin_border
#     ws.cell(row=k, column=1).border = thin_border
#     ws.cell(row=k, column=6).border = thin_border
#     ws.cell(row=k, column=7).border = thin_border
#     ws.cell(row=k, column=8).border = thin_border
#     ws.cell(row=k, column=9).border = thin_border
#     ws.cell(row=k, column=13).border = thin_border
#     k += 1
#     number += 1
#
# ws.merge_cells(start_row=11, start_column=2, end_column=2, end_row=k - 1)
# ws.merge_cells(start_row=11, start_column=3, end_column=3, end_row=k - 1)
# ws.merge_cells(start_row=11, start_column=4, end_column=4, end_row=k - 1)
# ws.merge_cells(start_row=11, start_column=5, end_column=5, end_row=k - 1)
#
# wb.save('Служебная записка на обработку и размножение чертежей.xlsx')

print(count_list)
print(ob_list)
print(naim_list)
print(naim)
print(naim2)
# print(naim3)
