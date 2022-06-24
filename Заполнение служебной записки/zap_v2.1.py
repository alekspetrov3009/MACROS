import os
import pythoncom
from win32com.client import Dispatch, gencache
from tkinter import Tk
from tkinter.filedialog import askopenfilenames
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
import datetime
import LDefin2D
import MiscellaneousHelpers as MH
from openpyxl.styles.numbers import BUILTIN_FORMATS

from tkinter import *
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import ttk



excel_path = os.getcwd()
shapka = ['№ п/п', '№ Сл.зап.', 'Дата', 'Тип тр-ра', '№ з/з', 'КТ', 'Дата', 'Инв. №', 'Чертежное\nобозначение', ' ',
          'Наименование', 'Кол-во листов', 'Формат']
k = 11  # номер строки
number = 1  # порядковый номер строки при заполнении
transformer_type = 'ЭОДЦН-8200/10-У3'   # Тип трансформатра

rukov = 'Уфрутов Р.С.'



if __name__ == "__main__":
    root = Tk()
    root.withdraw()  # Скрываем основное окно и сразу окно выбора файлов
    filenames = askopenfilenames(title="Выберите чертежи деталей",
                                 filetypes=[('Компас 3D', '*.cdw'), ('Компас 3D', '*.spw')])
    print(filenames)

    ispoln = simpledialog.askstring("SL Maker", "Введите фамилию и инициалы", parent=root)                # ФИО исполнителя
    order_number = simpledialog.askinteger("SL Maker", "Введите номер заказа", parent=root)                # Номер заказа
    slzap_number = simpledialog.askstring("SL Maker", "Введите номер служебной записки", parent=root)    # Порядковый номер служебной записки
    note_number = '52НН/'+ slzap_number + '/6-646'

root.destroy()  # Уничтожаем основное окно
root.mainloop()

print(ispoln)


excel_file = Workbook()
ws = excel_file.create_sheet(title='Служебка', index=0)

# Параметры стиля ячеек
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
# Параметры шрифта
font = Font(name='Times New Roman',
            size=12,
            color='FF000000',
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False)
title_font = Font(name='Times New Roman',
            size=22,
            color='FF000000',
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False)

# Параметры выравнивания
alignment = Alignment(horizontal='center',
                      vertical='center',
                      text_rotation=0,
                      wrap_text=True,
                      shrink_to_fit=False,
                      indent=0)

for i in filenames:
    #  Подключим описание интерфейсов API5
    kompas6_api5_module = gencache.EnsureModule("{0422828C-F174-495E-AC5D-D31014DBBE87}", 0, 1, 0)
    kompas_object = kompas6_api5_module.KompasObject(
        Dispatch("Kompas.Application.5")._oleobj_.QueryInterface(kompas6_api5_module.KompasObject.CLSID,
                                                                 pythoncom.IID_IDispatch))
    MH.iKompasObject = kompas_object

    #  Подключим описание интерфейсов API7
    kompas_api7_module = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0)
    application = kompas_api7_module.IApplication(
        Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(kompas_api7_module.IApplication.CLSID,
                                                                 pythoncom.IID_IDispatch))
    MH.iApplication = application
    MH.iApplication.Visible = False  # Видимость компаса

    Documents = application.Documents

    #  Открываем документ
    kompas_document = Documents.Open(i, False, False)
    kompas_document_2d = kompas_api7_module.IKompasDocument2D(kompas_document)
    iDocument2D = kompas_object.ActiveDocument2D()
    # Получаем свойства листов
    layout_sheets = kompas_document.LayoutSheets
    layout_sheet = layout_sheets.ItemByNumber(1)
    sheet_format = layout_sheet.Format
    sheet = kompas_document.LayoutSheets.Count  # счетчик листов
    # Считывание штампа
    iStamp = layout_sheet.Stamp
    iText = iStamp.Text(1)  # номер ячейки для считывания данных наименования
    naim = iText.Str
    naim = naim.replace('Сборочный чертеж', '')
    naim = naim.replace('\n', ' ')
    iText = iStamp.Text(2)  # номер ячейки для считывания данных обозначения
    oboz = iText.Str
    oboz = oboz.replace('БТЛИ.', '')
    iText = iStamp.Text(32)  # номер ячейки для считывания данных формата листа
    form = iText.Str
    # iStamp.ksCloseStamp()

    print(sheet)
    print(oboz)
    print(naim)
    print(form)

    # Ввод обозначения
    ws.cell(row=k, column=10).value = oboz
    ws.cell(row=k, column=10).border = thin_border
    ws.cell(row=k, column=10).font = font
    ws.cell(row=k, column=10).alignment = alignment
    # Ввод наименования
    ws.cell(row=k, column=11).value = naim
    ws.cell(row=k, column=11).border = thin_border
    ws.cell(row=k, column=11).font = font
    ws.cell(row=k, column=11).alignment = alignment
    # Ввод количества листов
    ws.cell(row=k, column=12).value = sheet
    ws.cell(row=k, column=12).font = font
    ws.cell(row=k, column=12).alignment = alignment
    ws.cell(row=k, column=12).border = thin_border
    # Ввод форматов листов
    ws.cell(row=k, column=13).value = form
    ws.cell(row=k, column=13).border = thin_border
    ws.cell(row=k, column=13).font = font
    ws.cell(row=k, column=13).alignment = alignment
    # Ввод порядкового номера
    ws.cell(row=k, column=1).value = number
    ws.cell(row=k, column=1).font = font
    ws.cell(row=k, column=1).alignment = alignment
    ws.cell(row=k, column=1).border = thin_border
    # Границы незадействованных ячеек
    ws.cell(row=k, column=6).border = thin_border
    ws.cell(row=k, column=7).border = thin_border
    ws.cell(row=k, column=8).border = thin_border
    ws.cell(row=k, column=13).border = thin_border
    # Ввод БТЛИ
    ws.cell(row=k, column=9).value = 'БТЛИ.'
    ws.cell(row=k, column=9).font = font
    ws.cell(row=k, column=9).alignment = alignment
    ws.cell(row=k, column=9).border = thin_border
    # Добавление строк
    ws.insert_rows(k + 1)
    # Увеличение строк
    k += 1
    number += 1
    # Ввод номеров столбцов
    for q in range(1, 14):
        ws.cell(row=10, column=q).value = q
        ws.cell(row=10, column=q).font = font
        ws.cell(row=10, column=q).alignment = alignment
        ws.cell(row=10, column=q).border = thin_border

    e = 1
    for i in shapka:
        ws.cell(row=9, column=e).value = i
        ws.cell(row=9, column=e).font = font
        ws.cell(row=9, column=e).alignment = alignment
        ws.cell(row=9, column=e).border = thin_border
        e += 1


# Объединить ячейки
def merge():
    ws.merge_cells(start_row=11, start_column=2, end_column=2, end_row=k - 1)
    ws.merge_cells(start_row=11, start_column=3, end_column=3, end_row=k - 1)
    ws.merge_cells(start_row=11, start_column=4, end_column=4, end_row=k - 1)
    ws.merge_cells(start_row=11, start_column=5, end_column=5, end_row=k - 1)
    # Объеденить ячейки 'чертежное обозначение'
    ws.merge_cells('I9:J9')
    ws.merge_cells('D4:J7')

#Изменение размеров ячеек
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 8
ws.column_dimensions['J'].width = 18
ws.column_dimensions['K'].width = 25
ws.column_dimensions['L'].width = 8
ws.column_dimensions['M'].width = 10


# Ввод шаблона

# Заведующая архивом
ws.cell(row=2, column=11).value = 'Зав. ЦА\nСамохиной А.А.'
ws.cell(row=2, column=11).font = font
ws.cell(row=2, column=11).alignment = alignment
ws['D4'].value = 'Служебная записка на обработку\nи размножение чертежей'
ws['D4'].font = title_font
ws['D4'].alignment = alignment

# Номер служебной записки
ws['B11'].value = note_number
ws['B11'].font = font
ws['B11'].alignment = alignment
ws['B11'].border = thin_border

# Текущая дата
# ws['C11'].value = date
ws['C11'].value = datetime.datetime.now()
ws['C11'].font = font
ws['C11'].alignment = alignment
ws['C11'].border = thin_border
ws['C11'].number_format = BUILTIN_FORMATS[14]

# Тип трансформатора
ws['D11'].value = transformer_type
ws['D11'].font = font
ws['D11'].alignment = alignment
ws['D11'].border = thin_border

# Номер заказа
ws['E11'].value = order_number
ws['E11'].font = font
ws['E11'].alignment = alignment
ws['E11'].border = thin_border

# Подписи внизу листа
ws.cell(row=k+5, column=4).value = 'Руководитель группы'
ws.cell(row=k+5, column=4).font = font
ws.cell(row=k+5, column=4).alignment = alignment

ws.cell(row=k+7, column=4).value = 'Исполнитель'
ws.cell(row=k+7, column=4).font = font
ws.cell(row=k+7, column=4).alignment = alignment

ws.cell(row=k+5, column=10).value = rukov
ws.cell(row=k+5, column=10).font = font
ws.cell(row=k+5, column=10).alignment = alignment

ws.cell(row=k+7, column=10).value = ispoln
ws.cell(row=k+7, column=10).font = font
ws.cell(row=k+7, column=10).alignment = alignment

merge()

# Закрытие фонового приложения Компаса
MH.iApplication.Quit()

# Сохранение эксель файла
excel_file.save(filename="Служебная записка на обработку и размножение чертежей.xlsx")

