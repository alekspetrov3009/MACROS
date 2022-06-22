import os
import re
import pythoncom
from win32com.client import Dispatch, gencache
from tkinter import Tk
from tkinter.filedialog import askopenfilenames
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
import LDefin2D
import MiscellaneousHelpers as MH

excel_path = os.getcwd()

if __name__ == "__main__":
    root = Tk()
    root.withdraw()  # Скрываем основное окно и сразу окно выбора файлов

    filenames = askopenfilenames(title="Выберите чертежи деталей", filetypes=[('Компас 3D', '*.cdw'), ])
    print(filenames)

    root.destroy()  # Уничтожаем основное окно
    root.mainloop()

k = 11      # номер строки
number = 1  # порядковый номер


for i in filenames:
    #  Подключим описание интерфейсов API5
    kompas6_api5_module = gencache.EnsureModule("{0422828C-F174-495E-AC5D-D31014DBBE87}", 0, 1, 0)
    kompas_object = kompas6_api5_module.KompasObject(
        Dispatch("Kompas.Application.5")._oleobj_.QueryInterface(kompas6_api5_module.KompasObject.CLSID,
                                                                 pythoncom.IID_IDispatch))
    MH.iKompasObject = kompas_object

    #  Подключим описание интерфейсов API7
    kompas_api7_module = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0)
    application = kompas_api7_module.IApplication(
        Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(kompas_api7_module.IApplication.CLSID,
                                                                 pythoncom.IID_IDispatch))
    MH.iApplication = application
    MH.iApplication.Visible = False  # Видимость компаса

    Documents = application.Documents

    #  Открываем документ
    kompas_document = Documents.Open(i, False, False)
    kompas_document_2d = kompas_api7_module.IKompasDocument2D(kompas_document)
    iDocument2D = kompas_object.ActiveDocument2D()
    # Получаем свойства листов
    layout_sheets = kompas_document.LayoutSheets
    layout_sheet = layout_sheets.ItemByNumber(1)
    sheet_format = layout_sheet.Format
    sheet = kompas_document.LayoutSheets.Count  # счетчик листов

    iStamp = layout_sheet.Stamp
    iText = iStamp.Text(1)  # номер ячейки для считывания данных наименования
    naim = iText.Str
    naim = naim.replace('Сборочный чертеж', '')
    iText = iStamp.Text(2)  # номер ячейки для считывания данных обозначения
    oboz = iText.Str
    oboz = oboz.replace('БТЛИ.', '')
    iText = iStamp.Text(32)  # номер ячейки для считывания данных формата листа
    form = iText.Str

    print(sheet)
    print(oboz)
    print(naim)
    print(form)

    wb = openpyxl.load_workbook("Служебная записка на обработку и размножение чертежей.xlsx")
    sheets = wb.sheetnames
    sheet_ranges = wb['Лист1']
    ws: Worksheet = wb.active
    ws.title = 'Лист1'

    # Параметры стиля ячеек
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # Параметры шрифта
    font = Font(name='Times New Roman',
                size=12,
                color='FF000000',
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False)

    # Параметры выравнивания
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)
# Ввод обозначения
    ws.cell(row=k, column=10).value = oboz
    ws.cell(row=k, column=10).border = thin_border
    ws.cell(row=k, column=10).font = font
    ws.cell(row=k, column=10).alignment = alignment
# Ввод наименования
    ws.cell(row=k, column=11).value = naim
    ws.cell(row=k, column=11).border = thin_border
    ws.cell(row=k, column=11).font = font
    ws.cell(row=k, column=11).alignment = alignment
# Ввод количества листов
    ws.cell(row=k, column=12).value = sheet
    ws.cell(row=k, column=12).font = font
    ws.cell(row=k, column=12).alignment = alignment
    ws.cell(row=k, column=12).border = thin_border
# Ввод форматов листов
    ws.cell(row=k, column=13).value = form
    ws.cell(row=k, column=13).border = thin_border
    ws.cell(row=k, column=13).font = font
    ws.cell(row=k, column=13).alignment = alignment
# Ввод порядкового номера
    ws.cell(row=k, column=1).value = number
    ws.cell(row=k, column=1).font = font
    ws.cell(row=k, column=1).alignment = alignment
    ws.cell(row=k, column=1).border = thin_border
# Границы незадействованных ячеек
    ws.cell(row=k, column=6).border = thin_border
    ws.cell(row=k, column=7).border = thin_border
    ws.cell(row=k, column=8).border = thin_border
    ws.cell(row=k, column=13).border = thin_border
# Ввод БТЛИ
    ws.cell(row=k, column=9).value = 'БТЛИ.'
    ws.cell(row=k, column=9).font = font
    ws.cell(row=k, column=9).alignment = alignment
    ws.cell(row=k, column=9).border = thin_border
# Добавление строк
    ws.insert_rows(k + 1)
# Увеличение строк
    k += 1
    number += 1

# Объединить ячейки
    def merge():
        ws.merge_cells(start_row=11, start_column=2, end_column=2, end_row=k-1)
        ws.merge_cells(start_row=11, start_column=3, end_column=3, end_row=k-1)
        ws.merge_cells(start_row=11, start_column=4, end_column=4, end_row=k-1)
        ws.merge_cells(start_row=11, start_column=5, end_column=5, end_row=k-1)

    merge()

    wb.save('Служебная записка на обработку и размножение чертежей.xlsx')
